import tkinter as tk
from tkinter import ttk
from tkinter import font
from tkinter import filedialog
import openpyxl

# Valores pré-definidos para os pesos
pesos_predefinidos = {
    "Administracao:": 0.8,
    "Classe de Alfabetização:": 0.2,
    "Creche:": 0.2,
    "Pré escola:": 0.2,
    "1 ano:": 0.5,
    "2 ano:": 0.5,
    "3 ano:": 0.5,
    "4 ano:": 0.5,
    "6 ano:": 0.5,
    "5 ano:": 0.5,
    "7 ano:": 0.5,
    "8 ano:": 0.5,
    "9 ano:": 0.5
}

# Variáveis globais para armazenar os valores dos pesos
pesos_values = pesos_predefinidos.copy()
pesos_entry = {}
pesos_window = None  # Definindo pesos_window globalmente

def definir_pesos():
    global pesos_entry
    global pesos_values
    global pesos_window  # Atribuindo a pesos_window globalmente
    pesos_window = tk.Toplevel(root)
    pesos_window.title("Definir Pesos")

    campos = [
        "Administracao:", "Classe de Alfabetização:", "Creche:", "Pré escola:", "1 ano:",
        "2 ano:", "3 ano:", "4 ano:", "5 ano:",
        "6 ano:", "7 ano:", "8 ano:", "9 ano:"
    ]

    pesos_entry = {}

    for i, campo in enumerate(campos):
        label = tk.Label(pesos_window, text=campo)
        label.grid(row=i, column=0, padx=5, pady=2, sticky="w")

        entry = tk.Entry(pesos_window)
        entry.grid(row=i, column=1, padx=5, pady=2)
        entry.insert(0, pesos_values.get(campo, "1.0"))
        pesos_entry[campo] = entry

    confirmar_button = tk.Button(pesos_window, text="Confirmar", command=confirmar_pesos)
    confirmar_button.grid(row=len(campos), columnspan=2, pady=10)

def confirmar_pesos():
    global pesos_values
    global pesos_entry
    global pesos_window  # Atribuindo a pesos_window globalmente

    for campo, entry in pesos_entry.items():
        pesos_values[campo] = float(entry.get())

    print("Pesos definidos:", pesos_values)

    pesos_window.destroy()

def inserir_caso():
    global entrada_entries
    entrada_window = tk.Toplevel(root)
    entrada_window.title("Inserir Caso de Entrada")
    entrada_entries = []

    campos = [
        "Administracao:","Classe de Alfabetização:", "Creche:", "Pré escola:", "1 ano:",
        "2 ano:", "3 ano:", "4 ano:", "5 ano:",
        "6 ano:", "7 ano:", "8 ano:", "9 ano:"
    ]

    for i, campo in enumerate(campos):
        label = tk.Label(entrada_window, text=campo)
        label.grid(row=i, column=0, padx=5, pady=2, sticky="w")

        if campo == "Administracao:":
            admin_combobox = ttk.Combobox(entrada_window, values=["Federal", "Estadual", "Municipal", "Particular"])
            admin_combobox.grid(row=i, column=1, padx=5, pady=2, sticky="we")
            entrada_entries.append(admin_combobox)
        else:
            entry = tk.Entry(entrada_window)
            entry.grid(row=i, column=1, padx=5, pady=2, sticky="we")
            entrada_entries.append(entry)

    gerar_similaridade_button = tk.Button(entrada_window, text="Gerar Similaridade", command=lambda: gerar_similaridade(entrada_window))
    gerar_similaridade_button.grid(row=len(campos), columnspan=2, pady=10)

def treeview_sort_column(tree, col, reverse=False):
    """Sorts a Treeview by a given column."""
    data = [(tree.set(child, col), child) for child in tree.get_children('')]
    data.sort(reverse=reverse)
    for index, (val, child) in enumerate(data):
        tree.move(child, '', index)

def gerar_similaridade(window):
    global pesos_values
    global entrada_entries
    count = 0

    # Lê o arquivo XLSX
    file_path = "Base_de_dados.xlsx"
    if file_path:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Criar Treeview para exibir as linhas da planilha em formato de tabela
        tree_frame = ttk.Frame(window)
        tree_frame.grid(row=len(entrada_entries) + 1, columnspan=2, padx=5, pady=5, sticky="nsew")

        # Criar um Canvas para conter a Treeview
        canvas = tk.Canvas(tree_frame)
        canvas.pack(side="top", fill="both", expand=True)

        # Adicionar uma barra de rolagem horizontal ao Canvas
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=canvas.xview)
        tree_scroll_x.pack(side="bottom", fill="x")

        tree = ttk.Treeview(tree_frame)
        tree.pack(side="left", fill="both", expand=True)

        # Configurar a barra de rolagem
        # tree_scroll = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
        # tree_scroll.pack(side="right", fill="y")
        # tree.configure(yscrollcommand=tree_scroll.set)

        # Configurar o Canvas para rolar horizontalmente com a barra de rolagem
        canvas.configure(xscrollcommand=tree_scroll_x.set)

        # Permitir a rolagem horizontal do Canvas
        def on_canvas_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        canvas.bind("<Configure>", on_canvas_configure)

        # Anexar a Treeview ao Canvas
        canvas.create_window((0, 0), window=tree, anchor="nw")

        # Permitir a rolagem do Canvas
        def on_tree_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))

        tree.bind("<Configure>", on_tree_configure)

        # Configurar as colunas
        tree["columns"] = sheet[1]
        tree.heading("#0", text="ID")
        tree.column("#0", width=50, stretch=False)
        tree.heading("#1", text="Ano")
        tree.heading("#2", text="Cód. Mun.")
        tree.heading("#3", text="Município")
        tree.heading("#4", text="Cód. INEP")
        tree.heading("#5", text="Nome da Escola")
        tree.heading("#6", text="Dep. Adm.")
        tree.heading("#7", text="Classe de Alfabetização")
        tree.heading("#8", text="Creche")
        tree.heading("#9", text="Pré escola")
        tree.heading("#10", text="1 ano")
        tree.heading("#11", text="2 ano")
        tree.heading("#12", text="3 ano")
        tree.heading("#13", text="4 ano")
        tree.heading("#14", text="5 ano")
        tree.heading("#15", text="6 ano")
        tree.heading("#16", text="7 ano")
        tree.heading("#17", text="8 ano")
        tree.heading("#18", text="9 ano")
        tree.heading("#19", text="Similaridade")

        for col, title in enumerate(sheet[1], start=1):
            tree.column(f"#{col}", width=100)  # Definindo largura padrão para as colunas

        # Adicionar linhas
        for row_data in sheet.iter_rows(min_row=3, values_only=True):

            if (count == 0):
                print(row_data)

            similaridade_adm = 0
            similaridade_classe_alfabetizacao = 0
            similaridade_creche = 0
            similaridade_pre_escola = 0
            similaridade_1_ano = 0
            similaridade_2_ano = 0
            similaridade_3_ano = 0
            similaridade_4_ano = 0
            similaridade_5_ano = 0
            similaridade_6_ano = 0
            similaridade_7_ano = 0
            similaridade_8_ano = 0
            similaridade_9_ano = 0
            similaridade_caso = 0
            soma_pesos = 0

            if 1:
                for column in range(1, len(row_data)):

                    #Similaridade de Administração
                    if column == 5:
                        if entrada_entries[0].get() == "Federal" and row_data[column] == "Federal":
                            similaridade_adm = 1
                        elif entrada_entries[0].get() == "Federal" and row_data[column] == "Estadual":
                            similaridade_adm = 0.6
                        elif entrada_entries[0].get() == "Federal" and row_data[column] == "Municipal":
                            similaridade_adm = 0.3
                        elif entrada_entries[0].get() == "Federal" and row_data[column] == "Particular":
                            similaridade_adm = 0
                        elif entrada_entries[0].get() == "Estadual" and row_data[column] == "Federal":
                            similaridade_adm = 0.6
                        elif entrada_entries[0].get() == "Estadual" and row_data[column] == "Estadual":
                            similaridade_adm = 1
                        elif entrada_entries[0].get() == "Estadual" and row_data[column] == "Municipal":
                            similaridade_adm = 0.6
                        elif entrada_entries[0].get() == "Estadual" and row_data[column] == "Particular":
                            similaridade_adm = 0
                        elif entrada_entries[0].get() == "Municipal" and row_data[column] == "Federal":
                            similaridade_adm = 0.3
                        elif entrada_entries[0].get() == "Municipal" and row_data[column] == "Estadual":
                            similaridade_adm = 0.6
                        elif entrada_entries[0].get() == "Municipal" and row_data[column] == "Municipal":
                            similaridade_adm = 1
                        elif entrada_entries[0].get() == "Municipal" and row_data[column] == "Particular":
                            similaridade_adm = 0
                        elif entrada_entries[0].get() == "Particular" and row_data[column] == "Federal":
                            similaridade_adm = 0
                        elif entrada_entries[0].get() == "Particular" and row_data[column] == "Estadual":
                            similaridade_adm = 0
                        elif entrada_entries[0].get() == "Particular" and row_data[column] == "Municipal":
                            similaridade_adm = 0
                        elif entrada_entries[0].get() == "Particular" and row_data[column] == "Particular":
                            similaridade_adm = 1

                    #Similaridade de Classe de Alfabetização
                    if column == 6:
                        classe_alfabetizacao_valor_1 = 0
                        classe_alfabetizacao_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] <= 50:
                            classe_alfabetizacao_valor_1 = 1
                        elif row_data[column] > 50 and row_data[column] <= 100:
                            classe_alfabetizacao_valor_1 = 2
                        elif row_data[column] > 100 and row_data[column] <= 200:
                            classe_alfabetizacao_valor_1 = 3
                        elif row_data[column] > 200 and row_data[column] < 300:
                            classe_alfabetizacao_valor_1 = 4
                        elif row_data[column] >= 300:
                            classe_alfabetizacao_valor_1 = 5

                        if float(entrada_entries[1].get()) >= 0 and float(entrada_entries[1].get()) <= 50:
                            classe_alfabetizacao_valor_2 = 1
                        elif float(entrada_entries[1].get()) > 50 and float(entrada_entries[1].get()) <= 100:
                            classe_alfabetizacao_valor_2 = 2
                        elif float(entrada_entries[1].get()) > 100 and float(entrada_entries[1].get()) <= 200:
                            classe_alfabetizacao_valor_2 = 3
                        elif float(entrada_entries[1].get()) > 200 and float(entrada_entries[1].get()) < 300:
                            classe_alfabetizacao_valor_2 = 4
                        elif float(entrada_entries[1].get()) >= 300:
                            classe_alfabetizacao_valor_2 = 5

                        similaridade_classe_alfabetizacao = 1 - ((abs(classe_alfabetizacao_valor_2 - classe_alfabetizacao_valor_1))/5)

                    #Similaridade de Creche
                    if column == 7:
                        creche_valor_1 = 0
                        creche_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 50:
                            creche_valor_1 = 1
                        elif row_data[column] >= 50 and row_data[column] < 100:
                            creche_valor_1 = 2
                        elif row_data[column] >= 100 and row_data[column] < 200:
                            creche_valor_1 = 3
                        elif row_data[column] >= 200 and row_data[column] < 300:
                            creche_valor_1 = 4
                        elif row_data[column] >= 300:
                            creche_valor_1 = 5

                        if float(entrada_entries[2].get()) >= 0 and float(entrada_entries[2].get()) < 50:
                            creche_valor_2 = 1
                        elif float(entrada_entries[2].get()) >= 50 and float(entrada_entries[2].get()) < 100:
                            creche_valor_2 = 2
                        elif float(entrada_entries[2].get()) >= 100 and float(entrada_entries[2].get()) < 200:
                            creche_valor_2 = 3
                        elif float(entrada_entries[2].get()) >= 200 and float (entrada_entries[2].get()) < 300:
                            creche_valor_2 = 4
                        elif float(entrada_entries[2].get()) >= 300:
                            creche_valor_2 = 5

                        similaridade_creche = 1 - ((abs(creche_valor_2 - creche_valor_1))/5)

                    #Similaridade de Pré escola
                    if column == 8:
                        pre_escola_valor_1 = 0
                        pre_escola_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 200:
                            pre_escola_valor_1 = 1
                        elif row_data[column] >= 200 and row_data[column] < 400:
                            pre_escola_valor_1 = 2
                        elif row_data[column] >= 400 and row_data[column] < 600:
                            pre_escola_valor_1 = 3
                        elif row_data[column] >= 600 and row_data[column] < 800:
                            pre_escola_valor_1 = 4
                        elif row_data[column] >= 800:
                            pre_escola_valor_1 = 5

                        if float(entrada_entries[3].get()) >= 0 and float(entrada_entries[3].get()) < 200:
                            pre_escola_valor_2 = 1
                        elif float(entrada_entries[3].get()) >= 200 and float(entrada_entries[3].get()) < 400:
                            pre_escola_valor_2 = 2
                        elif float(entrada_entries[3].get()) >= 400 and float(entrada_entries[3].get()) < 600:
                            pre_escola_valor_2 = 3
                        elif float(entrada_entries[3].get()) >= 600 and float(entrada_entries[3].get()) < 800:
                            pre_escola_valor_2 = 4
                        elif float(entrada_entries[3].get()) >= 800:
                            pre_escola_valor_2 = 5

                        similaridade_pre_escola = 1 - ((abs(pre_escola_valor_2 - pre_escola_valor_1))/5)

                    #Similaridade de 1 ano
                    if column == 9:
                        ano_1_valor_1 = 0
                        ano_1_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 50:
                            ano_1_valor_1 = 1
                        elif row_data[column] >= 50 and row_data[column] < 100:
                            ano_1_valor_1 = 2
                        elif row_data[column] >= 100 and row_data[column] < 150:
                            ano_1_valor_1 = 3
                        elif row_data[column] >= 150 and row_data[column] < 200:
                            ano_1_valor_1 = 4
                        elif row_data[column] >= 200:
                            ano_1_valor_1 = 5

                        if float(entrada_entries[4].get()) >= 0 and float(entrada_entries[4].get()) < 50:
                            ano_1_valor_2 = 1
                        elif float(entrada_entries[4].get()) >= 50 and float(entrada_entries[4].get()) < 100:
                            ano_1_valor_2 = 2
                        elif float(entrada_entries[4].get()) >= 100 and float(entrada_entries[4].get()) < 150:
                            ano_1_valor_2 = 3
                        elif float(entrada_entries[4].get()) >= 150 and float(entrada_entries[4].get()) < 200:
                            ano_1_valor_2 = 4
                        elif float(entrada_entries[4].get()) >= 200:
                            ano_1_valor_2 = 5

                        similaridade_1_ano = 1 - ((abs(ano_1_valor_2 - ano_1_valor_1))/5)

                    #Similaridade de 2 ano
                    if column == 10:
                        ano_2_valor_1 = 0
                        ano_2_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 100:
                            ano_2_valor_1 = 1
                        elif row_data[column] >= 100 and row_data[column] < 200:
                            ano_2_valor_1 = 2
                        elif row_data[column] >= 200 and row_data[column] < 300:
                            ano_2_valor_1 = 3
                        elif row_data[column] >= 300 and row_data[column] < 400:
                            ano_2_valor_1 = 4
                        elif row_data[column] >= 400:
                            ano_2_valor_1 = 5

                        if float(entrada_entries[5].get()) >= 0 and float(entrada_entries[5].get()) < 100:
                            ano_2_valor_2 = 1
                        elif float(entrada_entries[5].get()) >= 100 and float(entrada_entries[5].get()) < 200:
                            ano_2_valor_2 = 2
                        elif float(entrada_entries[5].get()) >= 200 and float(entrada_entries[5].get()) < 300:
                            ano_2_valor_2 = 3
                        elif float(entrada_entries[5].get()) >= 300 and float(entrada_entries[5].get()) < 400:
                            ano_2_valor_2 = 4
                        elif float(entrada_entries[5].get()) >= 400:
                            ano_2_valor_2 = 5

                        similaridade_2_ano = 1 - ((abs(ano_2_valor_2 - ano_2_valor_1))/5)

                    #Similaridade de 3 ano
                    if column == 11:
                        ano_3_valor_1 = 0
                        ano_3_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 100:
                            ano_3_valor_1 = 1
                        elif row_data[column] >= 100 and row_data[column] < 200:
                            ano_3_valor_1 = 2
                        elif row_data[column] >= 200 and row_data[column] < 300:
                            ano_3_valor_1 = 3
                        elif row_data[column] >= 300 and row_data[column] < 400:
                            ano_3_valor_1 = 4
                        elif row_data[column] >= 400:
                            ano_3_valor_1 = 5

                        if float(entrada_entries[6].get()) >= 0 and float(entrada_entries[6].get()) < 100:
                            ano_3_valor_2 = 1
                        elif float(entrada_entries[6].get()) >= 100 and float(entrada_entries[6].get()) < 200:
                            ano_3_valor_2 = 2
                        elif float(entrada_entries[6].get()) >= 200 and float(entrada_entries[6].get()) < 300:
                            ano_3_valor_2 = 3
                        elif float(entrada_entries[6].get()) >= 300 and float(entrada_entries[6].get()) < 400:
                            ano_3_valor_2 = 4
                        elif float(entrada_entries[6].get()) >= 400:
                            ano_3_valor_2 = 5

                        similaridade_3_ano = 1 - ((abs(ano_3_valor_2 - ano_3_valor_1))/5)

                    #Similaridade de 4 ano
                    if column == 12:
                        ano_4_valor_1 = 0
                        ano_4_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 100:
                            ano_4_valor_1 = 1
                        elif row_data[column] >= 100 and row_data[column] < 200:
                            ano_4_valor_1 = 2
                        elif row_data[column] >= 200 and row_data[column] < 300:
                            ano_4_valor_1 = 3
                        elif row_data[column] >= 300 and row_data[column] < 400:
                            ano_4_valor_1 = 4
                        elif row_data[column] >= 400:
                            ano_4_valor_1 = 5

                        if float(entrada_entries[7].get()) >= 0 and float(entrada_entries[7].get()) < 100:
                            ano_4_valor_2 = 1
                        elif float(entrada_entries[7].get()) >= 100 and float(entrada_entries[7].get()) < 200:
                            ano_4_valor_2 = 2
                        elif float(entrada_entries[7].get()) >= 200 and float(entrada_entries[7].get()) < 300:
                            ano_4_valor_2 = 3
                        elif float(entrada_entries[7].get()) >= 300 and float(entrada_entries[7].get()) < 400:
                            ano_4_valor_2 = 4
                        elif float(entrada_entries[7].get()) >= 400:
                            ano_4_valor_2 = 5

                        similaridade_4_ano = 1 - ((abs(ano_4_valor_2 - ano_4_valor_1))/5)

                    #Similaridade de 5 ano
                    if column == 13:
                        ano_5_valor_1 = 0
                        ano_5_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 100:
                            ano_5_valor_1 = 1
                        elif row_data[column] >= 100 and row_data[column] < 200:
                            ano_5_valor_1 = 2
                        elif row_data[column] >= 200 and row_data[column] < 300:
                            ano_5_valor_1 = 3
                        elif row_data[column] >= 300 and row_data[column] < 400:
                            ano_5_valor_1 = 4
                        elif row_data[column] >= 400:
                            ano_5_valor_1 = 5

                        if float(entrada_entries[8].get()) >= 0 and float(entrada_entries[8].get()) < 100:
                            ano_5_valor_2 = 1
                        elif float(entrada_entries[8].get()) >= 100 and float(entrada_entries[8].get()) < 200:
                            ano_5_valor_2 = 2
                        elif float(entrada_entries[8].get()) >= 200 and float(entrada_entries[8].get()) < 300:
                            ano_5_valor_2 = 3
                        elif float(entrada_entries[8].get()) >= 300 and float(entrada_entries[8].get()) < 400:
                            ano_5_valor_2 = 4
                        elif float(entrada_entries[8].get()) >= 400:
                            ano_5_valor_2 = 5

                        similaridade_5_ano = 1 - ((abs(ano_5_valor_2 - ano_5_valor_1))/5)

                    #Similaridade de 6 ano
                    if column == 14:
                        ano_6_valor_1 = 0
                        ano_6_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 150:
                            ano_6_valor_1 = 1
                        elif row_data[column] >= 150 and row_data[column] < 300:
                            ano_6_valor_1 = 2
                        elif row_data[column] >= 300 and row_data[column] < 450:
                            ano_6_valor_1 = 3
                        elif row_data[column] >= 450 and row_data[column] < 600:
                            ano_6_valor_1 = 4
                        elif row_data[column] >= 600:
                            ano_6_valor_1 = 5

                        if float(entrada_entries[9].get()) >= 0 and float(entrada_entries[9].get()) < 150:
                            ano_6_valor_2 = 1
                        elif float(entrada_entries[9].get()) >= 150 and float(entrada_entries[9].get()) < 300:
                            ano_6_valor_2 = 2
                        elif float(entrada_entries[9].get()) >= 300 and float(entrada_entries[9].get()) < 450:
                            ano_6_valor_2 = 3
                        elif float(entrada_entries[9].get()) >= 450 and float(entrada_entries[9].get()) < 600:
                            ano_6_valor_2 = 4
                        elif float(entrada_entries[9].get()) >= 600:
                            ano_6_valor_2 = 5

                        similaridade_6_ano = 1 - ((abs(ano_6_valor_2 - ano_6_valor_1))/5)

                    #Similaridade de 7 ano
                    if column == 15:
                        ano_7_valor_1 = 0
                        ano_7_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 150:
                            ano_7_valor_1 = 1
                        elif row_data[column] >= 150 and row_data[column] < 300:
                            ano_7_valor_1 = 2
                        elif row_data[column] >= 300 and row_data[column] < 450:
                            ano_7_valor_1 = 3
                        elif row_data[column] >= 450 and row_data[column] < 600:
                            ano_7_valor_1 = 4
                        elif row_data[column] >= 600:
                            ano_7_valor_1 = 5

                        if float(entrada_entries[10].get()) >= 0 and float(entrada_entries[10].get()) < 150:
                            ano_7_valor_2 = 1
                        elif float(entrada_entries[10].get()) >= 150 and float(entrada_entries[10].get()) < 300:
                            ano_7_valor_2 = 2
                        elif float(entrada_entries[10].get()) >= 300 and float(entrada_entries[10].get()) < 450:
                            ano_7_valor_2 = 3
                        elif float(entrada_entries[10].get()) >= 450 and float(entrada_entries[10].get()) < 600:
                            ano_7_valor_2 = 4
                        elif float(entrada_entries[10].get()) >= 600:
                            ano_7_valor_2 = 5

                        similaridade_7_ano = 1 - ((abs(ano_7_valor_2 - ano_7_valor_1))/5)

                    #Similaridade de 8 ano
                    if column == 16:
                        ano_8_valor_1 = 0
                        ano_8_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 100:
                            ano_8_valor_1 = 1
                        elif row_data[column] >= 100 and row_data[column] < 200:
                            ano_8_valor_1 = 2
                        elif row_data[column] >= 200 and row_data[column] < 300:
                            ano_8_valor_1 = 3
                        elif row_data[column] >= 300 and row_data[column] < 400:
                            ano_8_valor_1 = 4
                        elif row_data[column] >= 400:
                            ano_8_valor_1 = 5

                        if float(entrada_entries[11].get()) >= 0 and float(entrada_entries[11].get()) < 100:
                            ano_8_valor_2 = 1
                        elif float(entrada_entries[11].get()) >= 100 and float(entrada_entries[11].get()) < 200:
                            ano_8_valor_2 = 2
                        elif float(entrada_entries[11].get()) >= 200 and float(entrada_entries[11].get()) < 300:
                            ano_8_valor_2 = 3
                        elif float(entrada_entries[11].get()) >= 300 and float(entrada_entries[11].get()) < 400:
                            ano_8_valor_2 = 4
                        elif float(entrada_entries[11].get()) >= 400:
                            ano_8_valor_2 = 5

                        similaridade_8_ano = 1 - ((abs(ano_8_valor_2 - ano_8_valor_1))/5)

                    #Similaridade de 9 ano
                    if column == 17:
                        ano_9_valor_1 = 0
                        ano_9_valor_2 = 0
                        if row_data[column] >= 0 and row_data[column] < 150:
                            ano_9_valor_1 = 1
                        elif row_data[column] >= 150 and row_data[column] < 300:
                            ano_9_valor_1 = 2
                        elif row_data[column] >= 300 and row_data[column] < 450:
                            ano_9_valor_1 = 3
                        elif row_data[column] >= 450 and row_data[column] < 600:
                            ano_9_valor_1 = 4
                        elif row_data[column] >= 600:
                            ano_9_valor_1 = 5

                        if float(entrada_entries[12].get()) >= 0 and float(entrada_entries[12].get()) < 150:
                            ano_9_valor_2 = 1
                        elif float(entrada_entries[12].get()) >= 150 and float(entrada_entries[12].get()) < 300:
                            ano_9_valor_2 = 2
                        elif float(entrada_entries[12].get()) >= 300 and float(entrada_entries[12].get()) < 450:
                            ano_9_valor_2 = 3
                        elif float(entrada_entries[12].get()) >= 450 and float(entrada_entries[12].get()) < 600:
                            ano_9_valor_2 = 4
                        elif float(entrada_entries[12].get()) >= 600:
                            ano_9_valor_2 = 5

                        similaridade_9_ano = 1 - ((abs(ano_9_valor_2 - ano_9_valor_1))/5)

                if 1:
                    soma_pesos = float(pesos_values["Administracao:"]) + float(pesos_values["Classe de Alfabetização:"]) + float(pesos_values["Creche:"]) + float(pesos_values["Pré escola:"]) + float(pesos_values["1 ano:"]) + float(pesos_values["2 ano:"]) + float(pesos_values["3 ano:"]) + float(pesos_values["4 ano:"]) + float(pesos_values["5 ano:"]) + float(pesos_values["6 ano:"]) + float(pesos_values["7 ano:"]) + float(pesos_values["8 ano:"]) + float(pesos_values["9 ano:"])
                    similaridade_caso = similaridade_adm * float(pesos_values["Administracao:"]) + similaridade_classe_alfabetizacao * float(pesos_values["Classe de Alfabetização:"]) + similaridade_creche * float(pesos_values["Creche:"]) + similaridade_pre_escola * float(pesos_values["Pré escola:"]) + similaridade_1_ano * float(pesos_values["1 ano:"]) + similaridade_2_ano * float(pesos_values["2 ano:"]) + similaridade_3_ano * float(pesos_values["3 ano:"]) + similaridade_4_ano * float(pesos_values["4 ano:"]) + similaridade_5_ano * float(pesos_values["5 ano:"]) + similaridade_6_ano * float(pesos_values["6 ano:"]) + similaridade_7_ano * float(pesos_values["7 ano:"]) + similaridade_8_ano * float(pesos_values["8 ano:"]) + similaridade_9_ano * float(pesos_values["9 ano:"])
                    similaridade_caso = similaridade_caso / soma_pesos

                    if (count == 0):
                        print(similaridade_caso)

                    sheet.cell(row=count + 1, column=19, value=similaridade_caso)
                    tree.insert("", "end", text=count, values=row_data)

                #row_data[column] = pesos_values[sheet[1][column - 1].value] if row_data[column] == "S" else 0

            count += 1

        #treeview_sort_column(tree, '#19', reverse=True)
        workbook.save("Base_de_dados.xlsx")
        workbook.close()

        #tree.heading("#19", text="Similaridade", command=lambda: treeview_sort_column(tree, "#19", False))


root = tk.Tk()
root.title("Programa de Definição de Pesos e Inserção de Caso de Entrada")

botao_definir_pesos = tk.Button(root, text="Definir Pesos", command=definir_pesos)
botao_definir_pesos.pack(pady=10)

botao_inserir_caso = tk.Button(root, text="Inserir Caso de Entrada", command=inserir_caso)
botao_inserir_caso.pack(pady=10)

root.mainloop()
